#!/usr/bin/python

from Tkinter import *
from openpyxl import *
import tkMessageBox
from tkFileDialog import askopenfilename
from tkFileDialog import askdirectory
from operator import *
from datetime import datetime
from pymongo import MongoClient
from copy import *
from os import mkdir
import smtplib
from email.mime.text import MIMEText
from email.mime.application import MIMEApplication
from email.mime.multipart import MIMEMultipart
from smtplib import SMTP
from os import system
from os import popen
import random
from PIL import ImageTk
from PIL import Image
import tkFont
import re

def lalitfunction:
	print "testing for fun"

class Invigilator:
  def __init__(self,email="",name="",noOfExams=0,courses = []):
    self.email = email
    self.name = name
    self.noOfExams = noOfExams
    self.courses = courses

class Resource:
  def __init__(self,answerSheets= 0,graphPapers = 0,extraSheets = 0,tag=0):
    self.answerSheets = answerSheets
    self.graphPapers = graphPapers
    self.extraSheets = extraSheets
    self.tag = tag


class TimeTable :
  def __init__(self,examList=[],roomList=[]):
    self.examList = examList
    self.roomList = roomList

  def readTimeTable(self,ttFile):
    self.ttBook  = load_workbook(ttFile)
    self.ttSheet = self.ttBook.active
    self.noExams = self.ttSheet['A1'].value
    i = 0;
    self.examList=[]
    for i in range (0,self.noExams):
      rowNo = str(i+2)
      self.examCode = self.ttSheet['A'+rowNo].value
      self.examName = self.ttSheet['B'+rowNo].value
      self.examTime = self.ttSheet['C'+rowNo].value
      self.noOfStudents = self.ttSheet['D'+rowNo].value
      self.exam = Exam(courseTitle = self.examName, courseCode = self.examCode, examTime = self.examTime, noOfStudents = self.noOfStudents)
      self.examList.append(self.exam)
    for i in range (0,self.noExams):
      print self.examList[i].courseTitle,self.examList[i].courseCode,self.examList[i].examTime,self.examList[i].noOfStudents
    return 0

  def readRoomList(self,rlFile):
    self.rlBook  = load_workbook(rlFile)
    self.rlSheet = self.rlBook.active
    self.noRooms = self.rlSheet['A1'].value
    i = 0;
    self.roomList=[]
    for i in range (0,self.noRooms):
      rowNo = str(i+2)
      self.roomNo = self.rlSheet['A'+rowNo].value
      self.rows = self.rlSheet['B'+rowNo].value
      self.columns = self.rlSheet['C'+rowNo].value
      self.room = Room(roomNo = self.roomNo, rows = self.rows, columns = self.columns)
      self.roomList.append(self.room)
    for i in range (0,self.noRooms):
     print self.roomList[i].roomNo,self.roomList[i].rows,self.roomList[i].columns
    return 0

  def verifyTimeTable(self):
    self.capacity = 0
    self.totalStudents = 0
    for self.room in self.roomList: 
      self.capacity = self.capacity + (self.room.rows)*(self.room.columns)
    self.newList = sorted(self.examList, key=attrgetter('examTime'), reverse=False)
    self.examList = self.newList
    i = 0
    for i in range (0,self.noExams):
      print self.examList[i].courseTitle,self.examList[i].courseCode,self.examList[i].examTime,self.examList[i].noOfStudents
    self.prevTime = None
    self.slotStudents = 0
    self.valid = 1
    print "Total capacity is : ",self.capacity
    for self.exam in self.newList:
      if(self.exam.examTime == self.prevTime):
        self.slotStudents = self.slotStudents + self.exam.noOfStudents
        #print self.exam.examTime,self.slotStudents
      else:
        if(self.slotStudents > self.capacity):
          print " Error : At slot",self.prevTime,"the total students sitting for exam is",self.slotStudents,"which exceeds capacity",self.capacity
          self.valid = 0
        self.slotStudents = self.exam.noOfStudents
      self.prevTime = self.exam.examTime
    if(self.slotStudents > self.capacity):
      print " Error : At slot",self.prevTime,"the total students sitting for exam is",self.slotStudents,"which exceeds capacity",self.capacity
      self.valid = 0
    #print self.newList[1].examTime,self.newList[2].examTime;
    print "Time-table validity is : ",self.valid,"  ('1' if valid, '0' if invalid)"
    return self.valid

  def getTimeTable(self):
    return 0

class Exam:
  def __init__(self,courseTitle,courseCode,examTime,noOfStudents):
    self.courseTitle = courseTitle
    self.courseCode = courseCode
    self.examTime = examTime
    self.noOfStudents = noOfStudents

class Student:
  def __init__(self,name,rollNo,courseList,email):
    self.name = name
    self.rollNo = rollNo
    self.courseList = courseList  #List of course codes
    self.email = email

class AttendanceSheet:
  def __init__(self,roomNo,courseCode="",studentList=[]):
    self.courseCode = courseCode
    self.studentList = studentList
    self.roomNo = roomNo

class Course:
  def __init__(self,courseTitle="",courseCode="",studentList=[],instructor="",noOfStudents=0):
    self.courseTitle = courseTitle
    self.courseCode = courseCode
    self.studentList = studentList
    self.instructor = instructor
    self.noOfStudents = noOfStudents

class SeatingArrangement:
  def __init__(self,roomList=[],attendanceList=[],time=None):
    self.roomList = roomList
    self.attendanceList = attendanceList
    self.time = time

class Faculty:
  def __init__(self,name,email,courseList):
    self.name = name
    self.email = email
    self.courseList = courseList

  def getCourseList():
    return courseList

  def setCourseList(courseList):
    courseList = courseList

class Room:
  def __init__(self,roomNo,rows,columns,studentList = []):
    self.roomNo = roomNo
    self.rows = rows
    self.columns = columns
    if(len(studentList)!=0):
      self.studentList = studentList
    else:
      i = 0
      self.studentList=[]
      for i in range(0,self.columns):
        self.studentList.append([])

class Notification:
  def __init__(self):
    x = 0

  def notify(self):
    client = MongoClient()
    db = client.test
    cursor = db.makeup.find({},{'Email_ID':1,'_id':0,'Exam':1,'Name':1})
    server = smtplib.SMTP('smtp.gmail.com', 587)
    server.ehlo()
    server.starttls()
    #Next, log in to the server
    server.login("softtest360@gmail.com", "exammanage")

    #Send the mail
    for document in cursor:
      string1 = document["Exam"]
      msg = 'Subject: %s\n\n%s' % ("IMPORTANT", "Please contact your faculty ASAP. You have one make-up exam "+string1+" pending in our records.")
      server.sendmail("softtest360@gmail.com",document["Email_ID"], msg)
      print("DONE")
      self.studentName = document["Name"]
    tkMessageBox.showinfo("Complete", "Notified  all the entries")
    server.close()

class UI(Frame):
  def __init__(self,parent):
    Frame.__init__(self,parent)
    self.parent = parent
    print "Inside Example __init__"
    self.main2()
#    self.initUI()

  def initUI(self):
    self.frame.destroy()
    self.parent.title("Browse files " )
    self.frame=Frame(self.parent)
    fbg="white"
    self.frame1 = Frame(self.frame,background=fbg)
    self.frame2 = Frame(self.frame,background=fbg)
    self.frame3 = Frame(self.frame,background=fbg)
    self.frame4 = Frame(self.frame,background=fbg)

    self.fileSelectLBL = Label(self, text = "Please select Time-table file : ")
    self.fileSelectLBL.pack()
    print "Creating button now"
    abg="#f92727"
    mybg="orange"
    self.ttbutton = Button(self.frame1,text = "Browse Time-table", command = self.load_tt,width = 45,height = 5,activebackground=abg,bg=mybg,relief="groove")
    self.ttbutton.pack(side="left",pady=5)
    self.rlbutton = Button(self.frame2,text = "Browse Rooms-List", command = self.load_rl,width = 45,height = 5,activebackground=abg,bg=mybg,relief="groove")
    self.rlbutton.pack(side="left",pady=5)
    self.subbutton = Button(self.frame3,text = "Submit",pady=1, command = self.submit,width = 30,height = 5,activebackground=abg,bg=mybg,relief="groove")
    self.subbutton.pack(side="left",pady=25)
    self.back = Button(self.frame3,text = "Back", command = self.dest,width = 10,height = 3,activebackground=abg,bg=mybg,relief="groove")
    self.back.pack(side="bottom",pady=25)
    self.f1 = ""
    self.rl = ""
    self.frame.pack()
    self.frame1.pack(side='top',fill='both',expand=True)
    self.frame2.pack(side='top',fill='both',expand=True)
    self.frame3.pack(side='top',fill='both',expand=True)
    self.frame4.pack(side='bottom',fill='both',expand=True)

  def dest(self):
    self.frame.destroy()
    self.main2()

  def delete(self):
    abg = "#f92727"
    mybg="orange"
    fbg="white"
    self.g=0
    self.frame.destroy()
    self.initial=1
    self.parent.title("Exam Management Software " )
    self.frame = Frame(self.parent)
    self.frame1 = Frame(self.frame,background=fbg)
    self.frame2 = Frame(self.frame,background=fbg)
    self.frame3 = Frame(self.frame,background=fbg)
    self.frame31= Frame(self.frame,background=fbg)
    self.frame4 = Frame(self.frame,background=fbg)
    self.frame5 = Frame(self.frame,background=fbg)
    if(self.initial==1):
      self.find_roll1 = Label(self.frame1, text ="Roll No.",background = fbg)
      self.find_roll1.pack(expand=True,side=LEFT)
      self.find_roll2 = Entry(self.frame1,bd=5)
      self.find_roll2.pack(side=RIGHT)
      self.find_ = Button(self.frame2,text = "Find",pady=10,command=self.find,width = 10,height = 1,activebackground=abg,bg=mybg,relief="groove")
      self.find_.pack(expand=True,pady=5)
    self.sub = Button(self.frame31,text = "Back",pady=10,command=self.notifyMakeupExam,width = 8,height = 1,activebackground=abg,bg=mybg,relief="groove")
    self.sub.pack(expand=True,pady=8)
    self.frame.pack()
    self.frame1.pack(side='top',fill='both',expand=True)
    self.frame2.pack(side='top',fill='both',expand=True)
    self.frame3.pack(side='top',fill='both',expand=True)
    self.frame31.pack(side='top',fill='both',expand=True)
    self.frame4.pack(side='top',fill='both',expand=True)
    #self.frame5.pack(side='top',fill='both',expand=True)

  def del2(self):
    print("HERE2")
    client = MongoClient()
    db = client.test
    succ=0
    print(len(self.c))
    for i in range (len(self.c)):
      print(self.c[i].cget("text"))
      print (self.checkvar[i].get())
      if self.checkvar[i].get()==1:
        if((db.makeup.find({'Roll_No':self.find_roll2.get(),'Exam':self.c[i].cget("text")})).count!=0):
          tkMessageBox.showinfo("Success", "Removed entry/entries.")
          succ=1
        cursor = db.makeup.remove({'Roll_No':self.find_roll2.get().upper(),'Exam':self.c[i].cget("text")})
    if(succ==0):
      tkMessageBox.showinfo("Error", "Could not delete any record!")
    self.dest3()
       

  def find(self):
    self.frame31.destroy()
    fbg="white"
    mybg="orange"
    abg="#f92727"
    self.frame31= Frame(self.frame,background=fbg)
    self.checkvar=[]
    self.c=[]
    self.initial=1
    client = MongoClient()
    db = client.test
    cursor = db.makeup.find({'Roll_No':self.find_roll2.get().upper()})
    #if self.g>0:
    #  self.text.destroy()
    #self.text = Text(self.frame31,height=15,width=85)
    self.ij=0
    for document in cursor:
      self.g=1
      self.checkvar.append(IntVar())
      self.c.append(Checkbutton(self.frame31, text = "%s"%(document["Exam"]) , variable = self.checkvar[self.ij], \
                 onvalue = 1, offvalue = 0, height=5, \
                 width = 50))
      self.c[self.ij].pack()
      self.ij=self.ij+1
      #self.text.insert(INSERT,"Name : %s Email ID : %s Roll No. : %s Exam : %s \n"% (document["Name"],document["Email_ID"],document["Roll_No"],document["Exam"]))
      #self.text.pack()
    #if(self.initial==1):
     # self.find_roll3 = Label(self.frame3, text ="Roll No. to be deleted",background=fbg)
      #self.find_roll3.pack(expand=True,side=LEFT)
      
      #self.find_roll4 = Entry(self.frame3,bd=5)
      #self.find_roll4.pack(side=RIGHT)
    self.del21 = Button(self.frame31,text = "Delete Record",pady=10,command=self.del2,width = 18,height= 1,activebackground=abg,bg=mybg,relief="groove")
    self.del21.pack(expand=True)
    self.sub = Button(self.frame31,text = "Back",pady=10,command=self.notifyMakeupExam,width = 8,height = 1,activebackground=abg,bg=mybg,relief="groove")
    self.sub.pack(expand=True,pady=8)
    self.initial=0
    self.frame31.pack(side='top',fill='both',expand=True)
    #self.frame5.pack(side='top',fill='both',expand=True)

  def notifyMakeupExam(self):
    self.frame.destroy()
    self.parent.title("Makeup Manager " )
    abg = "#f92727"
    mybg="orange"
    fbg="white"
    self.frame = Frame(self.parent,background=fbg)
    self.frame1 = Frame(self.frame,background=fbg,highlightthickness=3,highlightbackground="#f93503")
    self.frame2 = Frame(self.frame,background=fbg,highlightthickness=3,highlightbackground="#f93503")
    self.frame3 = Frame(self.frame,background=fbg)
    self.frame4 = Frame(self.frame,background=fbg)
    self.frame5 = Frame(self.frame,background=fbg)
    notification = Notification()
    self.B = Button(self.frame1, text ="Notify About Makeup Exam",pady=5,command=notification.notify,width = 30,height = 3,activebackground=abg,bg=mybg,relief="groove")
    self.B.pack(expand=True,pady=5,side="top")
    self.C = Button(self.frame1,text = "See list of Makeup Exams",pady=5,command=self.see_list,width = 30,height = 3,activebackground=abg,bg=mybg,relief="groove")
    self.C.pack(expand=True,pady=5,side="bottom")
    self.D = Button(self.frame2,text = "Add Makeup Record",pady=5,command=self.add,width = 30,height = 3,activebackground=abg,bg=mybg,relief="groove")
    self.D.pack(expand=True,pady=5,side="top")
    self.initial=1
    self.E = Button(self.frame2, text ="Delete Record",pady=5,command=self.delete,width = 30,height = 3,activebackground=abg,bg=mybg,relief="groove")
    self.E.pack(expand= True,pady=5,side="bottom")
    self.F = Button(self.frame5,text = "Back",pady=5,command = self.dest1,width = 20,height = 2,activebackground=abg,bg=mybg,relief="groove")
    self.F.pack(expand=True,pady=10,side="bottom")
    #self.F.place(x = 50, y = 100, width=10, height=5)
    self.frame.pack()
    self.frame1.pack(side='top',fill='both',expand=True,ipadx=12)
    self.frame2.pack(side='top',fill='both',expand=True,pady=10)
    self.frame3.pack(side='top',fill='both',expand=True)
    self.frame4.pack(side='top',fill='both',expand=True)
    self.frame5.pack(side='bottom',fill='both',expand=True)

  def see_list(self):
    abg="#f92727"
    mybg="orange"
    self.frame.destroy()
    self.parent.title("Makeup students " )
    self.frame = Frame(self.parent,background="white")
    client = MongoClient()
    db = client.test
    cursor = db.makeup.find()
    text = Text(self.frame,width=100)
    for document in cursor:
      text.insert(INSERT,"Name : %s Email ID : %s Roll No. : %s Exam : %s \n"% (document["Name"],document["Email_ID"],document["Roll_No"],document["Exam"]))
      text.pack()
    self.sub = Button(self.frame,text = "Back",pady=10,command=self.notifyMakeupExam,width = 10,activebackground=abg,bg=mybg,relief="groove")
    self.sub.pack(expand=True)
    self.frame.pack(side='top',fill='both',expand=True)


  def add(self):
    self.frame.destroy()
    self.parent.title("Makeup Manager " )
    self.frame = Frame(self.parent)
    fbg= "white"
    abg="#f92727"
    mybg="orange"
    self.frame1 = Frame(self.frame,background=fbg)
    self.frame2 = Frame(self.frame,background=fbg)
    self.frame3 = Frame(self.frame,background=fbg)
    self.frame4 = Frame(self.frame,background=fbg)
    self.frame5 = Frame(self.frame,background=fbg)
    self.frame6 = Frame(self.frame,background=fbg)

    self.B1 = Label(self.frame1, text ="Name",bg=fbg)
    self.B1.pack(expand=True,side=LEFT)
    self.B2 = Entry(self.frame1,bd=5)
    self.B2.pack(side=RIGHT)
    self.C1 = Label(self.frame2, text ="Roll Number",bg=fbg)
    self.C1.pack(expand=True,side=LEFT)
    self.C2 = Entry(self.frame2,bd=5)
    self.C2.pack(side=RIGHT)
    self.D1 = Label(self.frame3, text ="Email ID",bg=fbg)
    self.D1.pack(expand=True,side=LEFT)
    self.D2 = Entry(self.frame3,bd=5)
    self.D2.pack(side=RIGHT)
    self.E1 = Label(self.frame4, text ="Exam-Name",bg=fbg)
    self.E1.pack(expand=True,side=LEFT)
    self.E2 = Entry(self.frame4,bd=5)
    self.E2.pack(side=RIGHT)
    self.sub = Button(self.frame5,text = "Submit",pady=10,command=self.addMakeup,width = 10,activebackground=abg,bg=mybg,relief="groove")
    self.sub.pack(expand=True,pady=5)
    self.sub = Button(self.frame6,text = "Back",pady=10,command=self.notifyMakeupExam,width = 10,activebackground=abg,bg=mybg,relief="groove")
    self.sub.pack(expand=True,pady=10)
    self.frame.pack()
    self.frame1.pack(side='top',fill='both',expand=True)
    self.frame2.pack(side='top',fill='both',expand=True)
    self.frame3.pack(side='top',fill='both',expand=True)
    self.frame4.pack(side='top',fill='both',expand=True)
    self.frame5.pack(side='top',fill='both',expand=True)
    self.frame6.pack(side='top',fill='both',expand=True)

  def addMakeup(self):
    client = MongoClient()
    db = client.test
    cursor = db.makeup.find({'Name':self.B2.get(),'Roll_No':self.C2.get().upper(),'Email_ID':self.D2.get(),'Exam':self.E2.get()})
    cursor2 = db.makeup.find({'Roll_No':self.C2.get().upper()})
    EMAIL_REGEX = re.compile(r"^[A-Za-z0-9\.\+_-]+@[A-Za-z0-9\._-]+\.[a-zA-Z]*$")
    if not self.B2.get() or not self.C2.get() or not self.D2.get() or not self.E2.get()  : 
      tkMessageBox.showinfo("Enter Again","Some fields were empty")
      self.add() 
    elif not EMAIL_REGEX.match(self.D2.get()):
      tkMessageBox.showinfo("Enter Again","Invalid E-Mail address")
      self.add()
    elif (cursor2.count())>0:
      self.gotIt = 0
      for docs in cursor2:
        if(docs["Name"]!=self.B2.get()):
          tkMessageBox.showinfo("Enter Again","Roll Number doesnt matched with the Name in the database")
          self.add()
          self.gotIt = 1
      if self.gotIt==0 and cursor.count()==0:
        result = db.makeup.insert({'Name':self.B2.get(),'Roll_No':self.C2.get().upper(),'Email_ID':self.D2.get(),'Exam':self.E2.get()})
        tkMessageBox.showinfo("Success", "Added entry.")
        self.dest3()
      elif cursor.count()!=0 :
        tkMessageBox.showinfo("Enter Again","Redundant entry")
        self.add()

    else :
      if cursor.count()!=1:
        #print self.B2.get(),self.C2.get(),self.D2.get(),self.E2.get()
        result = db.makeup.insert({'Name':self.B2.get(),'Roll_No':self.C2.get().upper(),'Email_ID':self.D2.get(),'Exam':self.E2.get()})
        tkMessageBox.showinfo("Success", "Added entry.")
        self.dest3()
      else:
        tkMessageBox.showinfo("Enter Again","Redundant entry")
        self.add()



  def dest3(self):
    self.frame.destroy()
    self.frame1.destroy()
    self.notifyMakeupExam()

  def main2(self):
    self.parent.title("Exam Management Software " )
    fbg = "white"
    self.frame = Frame(self.parent,background=fbg)
    self.frame1 = Frame(self.frame,background=fbg)
    self.frame2 = Frame(self.frame,background=fbg)
    self.frame3 = Frame(self.frame,background=fbg)
    self.frame4 = Frame(self.frame,background=fbg)
    abg = "#f92727"
    mybg="orange"
    self.B = Button(self.frame1, text ="Mail GuideLines",pady=10,command=self.sendExamGuidelines,width = 40,height = 4,activebackground=abg,bg=mybg,relief="sunken")
    self.C = Button(self.frame2,text = "Generate Seating Plan",pady=10,command=self.initUI,width = 40,height = 4,activebackground=abg,bg=mybg,relief="groove")
    self.C.pack(expand=True,pady = 4)
    self.B.pack(expand=True,pady=4)
    self.D = Button(self.frame3,text = "Makeup Manager",pady=10,command=self.notifyMakeupExam,width = 40,height = 4,activebackground=abg,bg=mybg,relief="groove")
    self.D.pack(expand=True,pady=4)
    self.E = Button(self.frame4, text ="Help",pady=10,command=self.help,width = 40,height = 3,activebackground=abg,bg=mybg,relief="sunken")
    self.E.pack(expand= True,pady=4)
    self.frame.pack()
    self.frame2.pack(side='top',fill='both',expand=True)
    self.frame1.pack(side='top',fill='both',expand=True)
    self.frame3.pack(side='top',fill='both',expand=True)
    self.frame4.pack(side='bottom',fill='both',expand=True)

  def help(self):
    popen("evince file:"+"examhelp.pdf")

  def sendExamGuidelines(self):
    self.frame.destroy()
    self.parent.title("Mailing Guidelines " )
    self.customFont = tkFont.Font(family="Helvetica", size=18)
    abg = "#f92727"
    mybg="orange"
    self.frame = Frame(self.parent,background="white")
    self.frame1 = Frame(self.frame,background="white")
    self.guideFile=""
    self.label = Label(self.frame,text = "Please select the guidelines file to be sent to all students.",font=self.customFont,background="white")
    self.label.pack(expand=True,pady=10)
    self.mail1 = Button(self.frame,text = "Browse the guidelines file",pady=10,command=self.browsePdf,width = 40,height = 4,activebackground=abg,bg=mybg,relief="groove")
    self.mail1.pack(expand=True,pady=10)
    self.mail = Button(self.frame,text = "Mail Guidelines",pady=10,command=self.mail,width = 40,height = 4,activebackground=abg,bg=mybg,relief="groove")
    self.mail.pack(expand=True,pady=10)
    self.back = Button(self.frame1,text = "Back",pady=10,command=self.dest1,width = 20,height = 2,activebackground=abg,bg=mybg,relief="groove")
    self.back.pack(expand=True,pady=40,side="left")
    self.frame.pack()
    self.frame1.pack()

  def dest1(self):
    self.frame.destroy()
    self.frame1.destroy()
    self.main2()

  def browsePdf(self):
    ftypes = [("PDF Files","*.pdf")]
    self.guideFile = askopenfilename(filetypes = ftypes)    
    
  def mail(self):
    if(self.guideFile==""):
      tkMessageBox.showinfo("Error", "Please select guidelines file")
    else:
      self.msg = MIMEMultipart()
      self.msg['Subject'] = 'Important: Exam Guiedlines'
      self.msg['From'] = 'softtest360@gmail.com'
      #msg['Reply-to'] = 'otroemail@dominio'
      self.msg['To'] = 'paliwal.2@iitj.ac.in'
 
      # That is what u see if dont have an email reader:
      self.msg.preamble = 'Multipart massage.\n'
 
      # This is the textual part:
      self.part = MIMEText("Please find attched guidelines for exams.")
      self.msg.attach(self.part)
 
      # This is the binary part(The Attachment):
      self.part = MIMEApplication(open(self.guideFile,'rb').read())
      self.part.add_header('Content-Disposition', 'attachment', filename=self.guideFile)
      self.msg.attach(self.part)
 
      # Create an instance in SMTP server
      self.server = SMTP("smtp.gmail.com",587)
      # Start the server:
      self.server.ehlo()
      self.server.starttls()
      self.server.login("softtest360@gmail.com", "exammanage")
 
    # Send the email
      self.server.sendmail(self.msg['From'], self.msg['To'], self.msg.as_string())
      tkMessageBox.showinfo("Success", "Guidelines have been successfully mailed.")
      self.dest1()


  def load_tt(self, ftypes = None):
    ftypes = [("Excel files","*.xlsx")]
    self.f1 = askopenfilename(filetypes = ftypes)

  def load_rl(self, ftypes = None):
    ftypes = [("Excel files","*.xlsx")]
    self.rl = askopenfilename(filetypes = ftypes)

  def load_studFile(self,ftypes = None):
    ftypes = [("Excel files","*.xlsx")]
    self.studFile = askopenfilename(filetypes = ftypes)

  def load_instrFile(self,ftypes = None):
    ftypes = [("Excel files","*.xlsx")]
    self.instrFile = askopenfilename(filetypes = ftypes)

  def load_output(self,ftypes=None):
    self.out_path = askdirectory(mustexist=1)

  def submit(self):
    #lbl = Label(self.parent,text="Please Browse Files: ")
    #lbl.pack()
    print self.f1
    print self.rl
    if ((self.f1=="") or (self.rl=="")):
      tkMessageBox.showinfo("Error", "No input files provided")
    else :
     # lbl.config(text="Getting data")
      self.tt = TimeTable()
      try:
        self.tt.readTimeTable(ttFile = self.f1)
      except:
        tkMessageBox.showinfo("Error", "Invalid Time-table file format!")
	return 0
      try:
        self.tt.readRoomList(rlFile = self.rl)
      except:
        tkMessageBox.showinfo("Error", "Invalid rooms list file format!")
        return 0
      try:
        validity = self.tt.verifyTimeTable()
      except:
        tkMessageBox.showinfo("Error", "Unable to verify time-table! File-formats or entries may be incorrect!")
        return 0
      if validity==0 :
        tkMessageBox.showinfo("Sorry", "Time-table is not valid! The rooms capacity is less than the total no of students for exam in a slot!        PLEASE GIVE CORRECT TIME-TABLE")
        return 0
      self.frame.destroy()
      self.parent.title("Some more files please: " )
      fbg="white"
      abg="#f92727"
      mybg="orange"
      self.out_path=""
      self.frame = Frame(self.parent,background = fbg)
      self.B = Button(self.frame, text ="Browse Students List File: ",pady=10,command = self.load_studFile,width = 40,height = 1,activebackground=abg,bg=mybg,relief="groove")
      self.C = Button(self.frame, text ="Browse Instructors List ",pady=10,command = self.load_instrFile,width = 40,height = 1,activebackground=abg,bg=mybg,relief="groove")
      self.D = Button(self.frame, text ="Browse Output Destination",pady=10,command = self.load_output,width = 40,height = 1,activebackground=abg,bg=mybg,relief="groove")
      self.E = Button(self.frame, text ="Submit", pady=10,command = self.generateAll,width = 25,height = 1,activebackground=abg,bg=mybg,relief="groove")
      self.F = Button(self.frame, text ="Back to Home",pady=10,command=self.dest1,width = 15,height = 1,activebackground=abg,bg=mybg,relief="groove")
      self.B.pack(expand=True,pady=3)
      self.C.pack(expand=True,pady=3)
      self.D.pack(expand=True,pady=3)
      self.E.pack(expand=True,pady=10)
      self.F.pack(expand=True,pady=30)
      self.frame.pack()
    self.studFile = ""
    self.instrFile = ""
    print "Got !"

  def generateAll(self):
    try:
      print self.studFile
      print self.instrFile
      print self.out_path
      if ((self.studFile=="") or (self.instrFile=="") or (self.out_path=="")):
        tkMessageBox.showinfo("Error", "Please Provide all inputs")
      else:
        self.studBook  = load_workbook(self.studFile)
        self.studSheet = self.studBook.active
        self.noStudents = self.studSheet['A1'].value
        i = 0
        self.studentList = []
        self.courses = []
        curr=1
        for i in range (0,self.noStudents):
          curr=curr+1
          rowNo = str(curr)
          self.name = self.studSheet['A'+rowNo].value
          self.rollNo = self.studSheet['B'+rowNo].value
          self.email = self.studSheet['C'+rowNo].value
          self.noOfCourses = self.studSheet['D'+rowNo].value
          j = 0
          self.courseList = []
          for j in range (0,self.noOfCourses):
            curr = curr+1
            rowNo = str(curr)
            self.courseList.append(self.studSheet['A'+rowNo].value)
            found = 0
            self.corCode = self.studSheet['A'+rowNo].value
            self.corValue = self.studSheet['B'+rowNo].value
            for self.item in self.courses:
              if (self.item.courseCode == self.corCode):
                found = 1
                self.item.studentList.append(self.rollNo)
                self.item.noOfStudents = self.item.noOfStudents+1
            if (found == 0):
              self.courses.append(Course(courseCode = self.corCode,courseTitle = self.corValue,studentList = [self.rollNo],noOfStudents=1))
            found = 0
          self.student = Student(name = self.name,rollNo = self.rollNo,email=self.email,courseList = self.courseList)
          self.studentList.append(self.student)
          stud = None
        for stud in self.studentList:
          print stud.name,stud.rollNo,stud.email
          for self.course in stud.courseList:
            print self.course
        stri = None
        for stud in self.courses:
          print stud.courseCode,stud.courseTitle,stud.noOfStudents
          for stri in stud.studentList:
            print stri
        self.generateSeatingArrangement();
        return 0
    except:
      tkMessageBox.showinfo("Error", "Invalid format of input files")

  def generateSeatingArrangement(self):
    stud = None
    #self.courses = sorted(self.courses, key=attrgetter('noOfStudents'), reverse=True)
    ## Here is our algorithm, working on processed data
    #self.row1 = None
    #self.tt.roomList[1].studentList.append(self.row1)
    self.roomList = self.tt.roomList
    self.salist = []
    self.examList = []
    self.prevTime = None
    #self.examPr = None
    for self.exam in self.tt.examList:
      if(self.exam.examTime==self.prevTime):
        self.examList.append(self.exam)
      else:
        ###### ALGORITHM SHOULD BE WORKING IN THIS PART
        if (len(self.examList)==0):
          self.examList.append(self.exam)
          self.prevTime = self.exam.examTime
          continue
        k=0
        self.roomList = []
        self.roomList = deepcopy(self.tt.roomList)
        for self.examPr in self.examList:
          self.noOfStudents = self.examPr.noOfStudents
          print "Inside for ",self.noOfStudents,self.examPr.courseTitle,self.examPr.courseCode
          self.courseCode = self.examPr.courseCode
          self.studentList1=[]
          for self.cour in self.courses:
            if self.cour.courseCode == self.courseCode:
              self.studentList1 = self.cour.studentList
              break
          count = 0
          for self.room in self.roomList:
            col = 1
            var = 0
            self.row1 = []
            j=0
            for j in range (0,len(self.room.studentList)):
              #print self.row1
              if(col%2==1):
                if(len(self.room.studentList[j])!=self.room.rows):
                  for var in range (len(self.room.studentList[j]),self.room.rows):
                    if(count < len(self.studentList1)):
                      print 'count= ',count
                      self.room.studentList[j].append(self.studentList1[count])
                      count = count + 1
                    else:
                      break
              col = col+1
            if(count >= len(self.studentList1)):
              break
          for self.room in self.roomList:
              col = 1
              var = 0
              self.row1 = []
              j=0
              for j in range (0,len(self.room.studentList)):
                #print self.row1
                if(col%2==0):
                  if(len(self.room.studentList[j])!=self.room.rows):
                    for var in range (len(self.room.studentList[j]),self.room.rows):
                      if(count < len(self.studentList1)):
                        print 'count= ',count
                        self.room.studentList[j].append(self.studentList1[count])
                        count = count + 1
                      else:
                        break
                col = col+1
              if(count >= len(self.studentList1)):
                break
        #Here we add a sa object
        self.sa = SeatingArrangement(roomList = self.roomList,time = self.prevTime)
        self.salist.append(self.sa)
        ###### ALGORITHM PART ENDS HERE
        self.examList = []
        self.examList.append(self.exam)
        self.prevTime = self.exam.examTime

    ###### ALGORITHM SHOULD BE WORKING IN THIS PART
    k=0
    self.roomList = []
    self.roomList = deepcopy(self.tt.roomList)
    for self.examPr in self.examList:
      self.noOfStudents = self.examPr.noOfStudents
      print "Inside for ",self.noOfStudents,self.examPr.courseTitle,self.examPr.courseCode
      self.courseCode = self.examPr.courseCode
      self.studentList1=[]
      for self.cour in self.courses:
        if self.cour.courseCode == self.courseCode:
          self.studentList1 = self.cour.studentList
          break
      count = 0
      for self.room in self.roomList:
        col = 1
        var = 0
        self.row1 = []
        j=0
        for j in range (0,len(self.room.studentList)):
          #print self.row1
          if(col%2==1):
            if(len(self.room.studentList[j])!=self.room.rows):
              for var in range (len(self.room.studentList[j]),self.room.rows):
                if(count < len(self.studentList1)):
                  print 'count= ',count
                  self.room.studentList[j].append(self.studentList1[count])
                  count = count + 1
                else:
                  break
          col = col+1
        if(count >= len(self.studentList1)):
          break
      for self.room in self.roomList:
        col = 1
        var = 0
        self.row1 = []
        j=0
        for j in range (0,len(self.room.studentList)):
          #print self.row1
          if(col%2==0):
            if(len(self.room.studentList[j])!=self.room.rows):
              for var in range (len(self.room.studentList[j]),self.room.rows):
                if(count < len(self.studentList1)):
                  print 'count= ',count
                  self.room.studentList[j].append(self.studentList1[count])
                  count = count + 1
                else:
                  break
          col = col+1
        if(count >= len(self.studentList1)):
          break

    #Here we add a sa object
    self.sa = SeatingArrangement(roomList = self.roomList,time = self.prevTime)
    self.salist.append(self.sa)
    ###### ALGORITHM PART ENDS HERE
    self.examList = []
    self.examList.append(self.exam)
    self.prevTime = self.exam.examTime
    for self.sa in self.salist:
      self.instBook = load_workbook(self.instrFile)
      self.instSheet = self.instBook.active
      self.noOfInvigilators = self.instSheet['A1'].value
      self.invigList=[]
      self.allocList=[]
      variable=0
      for variable in range(0,self.noOfInvigilators):
        self.invigList.append(self.instSheet['A'+str(variable+2)].value)
        self.allocList.append(0)
      self.avgSize = int(self.noOfInvigilators/len(self.sa.roomList))
      if(self.avgSize>=2):
        self.avgSize = 2
      if(self.avgSize<1):
        tkMessageBox.showinfo("Error", "Insufficient no. of invigilators. Please update the Invigilators list file")
      else:
        i = 0
        p=0
        invigwb = Workbook()
        invigws = invigwb.active
        invigws['A1'] = "Invigilator Duty for "+self.sa.time
        rowNo = 2
        listAlpha = ['B','C']
        for p in range(0,len(self.sa.roomList)):
          self.l1=[]
          print self.avgSize
          invigws['A'+str(rowNo)] = self.sa.roomList[p].roomNo
          for i in range(0,self.avgSize):
            k = random.randint(0,self.noOfInvigilators-1)
            while (self.allocList[k]==1):
              k = random.randint(0,self.noOfInvigilators-1)
            invigws[listAlpha[i]+str(rowNo)] = self.invigList[k]
            print self.invigList[k],k
            self.allocList[k] = 1
          rowNo = rowNo + 1
      
        try:
          self.path = self.out_path
          self.path = self.path+"/SA_"+self.sa.time
          mkdir(self.path,0744)
        except OSError:
          pass
        invigwb.save(self.path+"/Invigilator_duty.xlsx")
      listAlpha = ['A','B','C','D','E','F','G','H','I','J','K','L','M','N','O','P','Q','R','S','T','U','V','W','X','Y','Z']   #Limits no. of columns to 26
      self.examList = []
      for self.exam in self.tt.examList:
        if(self.exam.examTime == self.sa.time):
          self.examList.append(self.exam)
          #print self.exam.courseCode
      for self.room in self.sa.roomList:
        colNo = 0
        wb  = Workbook()
        ws = wb.active
        self.attendanceLists = []
        ws['A1'] = str(self.sa.time)+" Room No. " +str(self.room.roomNo)
        for self.row1 in self.room.studentList:
          rowNo = 2
          for self.student in self.row1:
            ws[listAlpha[colNo]+str(rowNo)] = self.student
            rowNo = rowNo + 1
            got = 0
            #print "Will locate student: ",self.student
            for self.exam in self.examList:
              #print "Loop1 ",self.exam.courseCode
              for self.cour in self.courses:
                #print "Loop2 ",self.cour.courseCode
                if self.cour.courseCode == self.exam.courseCode:
                  for self.stud in self.cour.studentList:
                    #print "Loop3, searching for : ",self.student," found : ",self.stud
                    if self.stud == self.student:
                      succ = 0
                      for self.attList in self.attendanceLists:
                        #print "Loop4, searching for ",self.cour.courseCode," found: ",self.attList.courseCode
                        if (self.attList.courseCode == self.cour.courseCode):
                          self.attList.studentList.append(self.student)
                          succ=1
                          got = 1
                          break
                      if (succ==0):
                        #print "Creating for ",self.cour.courseCode
                        self.atten = AttendanceSheet(roomNo = self.room.roomNo,courseCode=self.cour.courseCode,studentList=[self.student])
                        self.attendanceLists.append(self.atten)
                        got = 1
                        break
                    if(got==1):
                      break
                if(got == 1):
                  break
              if(got==1):
                break
          colNo = colNo + 1
        ws.column_dimensions["A"].width = 40.0
        ws.row_dimensions[1].height = 20
        fontObj2 = styles.Font(size=16, italic=True)
        styleObj2 = styles.Style(font=fontObj2)
        ws['A1'].style = styleObj2
        wb.save(self.path+'/SA_'+str(self.room.roomNo)+".xlsx")

        print "Size of list: ",len(self.attendanceLists)
        print "Room No.: ",self.room.roomNo
        for self.att in self.attendanceLists:
          self.path = self.out_path
          self.path = self.path+"/SA_"+self.sa.time
          try:
            mkdir(self.path+"/Attendance_Sheets_RoomNo"+str(self.att.roomNo))
          except OSError:
            pass
          self.wb = Workbook()
          self.ws = self.wb.active
          self.ws['A1']= 'Attendance Sheet '+str(self.att.courseCode)+"Room No.:"+str(self.room.roomNo)
          self.ws['A2']="Name"
          self.ws['B2']="Signature"
          row = 3
          for self.stu in self.att.studentList:
            for self.st in self.studentList:
              if (self.st.rollNo == self.stu):
                self.stu = self.st.name+"( "+self.stu+" )"
                break
            self.ws['A'+str(row)] = self.stu
            row = row+1
          self.ws.column_dimensions["A"].width = 40.0
          self.ws.row_dimensions[1].height = 20
          fontObj2 = styles.Font(size=16, italic=True)
          styleObj2 = styles.Style(font=fontObj2)
          self.ws['A1'].style = styleObj2
          self.wb.save(self.path+"/Attendance_Sheets_RoomNo"+str(self.att.roomNo)+"/Attendance_Sheet_"+str(self.att.courseCode)+".xlsx")
          print self.att.courseCode
          for self.sRoll in self.att.studentList:
            print self.sRoll
    tkMessageBox.showinfo("Success", "All Ouptput files are generated in "+self.out_path+" folder.")
    self.dest1()

def main(): 
  window = Tk()
  canvas = Canvas(width = 200, height = 170, bg = 'blue')
  canvas.pack(expand = NO, fill = BOTH)

  image = ImageTk.PhotoImage(file = "BG.png")
  canvas.create_image(0, 0, image = image, anchor = NW)
  print "inside main"
  uiObject = UI(window)
  fbg="white"
  window.configure(background=fbg)
  window.geometry("800x600")
  window.title("Exam Management Software")
  window.mainloop()

main()
